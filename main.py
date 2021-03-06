import os
import sqlite3
from openpyxl import load_workbook
from random import randrange
import platform
import subprocess
import ctypes


def import_data(handler, platform):
    # ファイルを読み込む
    if platform == "Kahoot!":
        # Kahoot!
        filename = "kahoot.xlsx"
    elif platform == "Quizizz":
        # Quizizz
        filename = "quizizz.xlsx"
    else:
        # サポートしないプラットフォーム
        msgbox(platform + " に対応していません。", "サポートしないプラットフォーム")
        exit(1)
    # データを取り込む
    vocabulary = []
    wb = load_workbook("input/" + filename)
    ws = wb.worksheets[0]
    for row in ws.rows:
        if row[0].value is None:
            # 空白セル（単語列）
            msgbox(
                "input/" + filename + " の " + row[0].coordinate + " に内容を入力してください。",
                "空白セル（単語列）",
            )
            exit(1)
        if row[1].value is None:
            # 空白セル（意味列）
            msgbox(
                "input/" + filename + " の " + row[1].coordinate + " に内容を入力してください。",
                "空白セル（意味列）",
            )
            exit(1)
        vocabulary.append((row[0].value, row[1].value))
    # データベースにインポートする
    handler.executemany(
        "INSERT INTO vocabulary(word, meaning) VALUES (?, ?)", vocabulary
    )


def generate_upload_file(handler, platform):
    if platform == "Kahoot!":
        # Kahoot!
        generate_kahoot(handler)
        os.rename("input/kahoot.xlsx", "input/_kahoot.xlsx")
    elif platform == "Quizizz":
        # Quizizz
        generate_quizizz(handler)
        os.rename("input/quizizz.xlsx", "input/_quizizz.xlsx")
    else:
        # サポートしないプラットフォーム
        msgbox(platform + " に対応していません。", "サポートしないプラットフォーム")
        exit(1)


def generate_kahoot(handler):
    # Kahoot!
    wb = load_workbook("template/kahoot.xlsx")
    ws = wb.worksheets[0]
    curRow = 9
    # フォーマット定義
    colQuestion = "B"
    colOption1 = "C"
    colOption2 = "D"
    colOption3 = "E"
    colOption4 = "F"
    colTimeLimit = "G"
    colAnswer = "H"
    # データを抽出する
    questions = handler.execute(
        "SELECT * FROM vocabulary ORDER BY RANDOM() LIMIT 20"
    ).fetchall()
    for question in questions:
        questionId = question[0]
        questionWord = question[1]
        questionMeaning = question[2]
        # 問題文を作成する
        ws[colQuestion + str(curRow)] = questionWord
        # 選択肢を作成する
        options = handler.execute(
            "SELECT meaning FROM vocabulary WHERE id <> "
            + str(questionId)
            + " ORDER BY RANDOM() LIMIT 4"
        ).fetchall()
        ws[colOption1 + str(curRow)] = options[0][0]
        ws[colOption2 + str(curRow)] = options[1][0]
        ws[colOption3 + str(curRow)] = options[2][0]
        ws[colOption4 + str(curRow)] = options[3][0]
        # 時間制限を指定する
        ws[colTimeLimit + str(curRow)] = 5
        # 1/2/3/4の中で正解を決める
        answerOption = randrange(1, 5)
        ws[colAnswer + str(curRow)] = answerOption
        if answerOption == 1:
            ws[colOption1 + str(curRow)] = questionMeaning
        elif answerOption == 2:
            ws[colOption2 + str(curRow)] = questionMeaning
        elif answerOption == 3:
            ws[colOption3 + str(curRow)] = questionMeaning
        elif answerOption == 4:
            ws[colOption4 + str(curRow)] = questionMeaning
        else:
            # 何もしない
            pass
        curRow += 1
    # ファイルを保存する
    wb.save("output/kahoot.xlsx")


def generate_quizizz(handler):
    # Quizizz
    wb = load_workbook("template/quizizz.xlsx")
    ws = wb.worksheets[0]
    curRow = 3
    # フォーマット定義
    colQuestion = "A"
    colQuestionType = "B"
    colOption1 = "C"
    colOption2 = "D"
    colOption3 = "E"
    colOption4 = "F"
    colOption5 = "G"
    colAnswer = "H"
    colTimeLimit = "I"
    colImageLink = "J"
    # データを抽出する
    questions = handler.execute(
        "SELECT * FROM vocabulary ORDER BY RANDOM() LIMIT 20"
    ).fetchall()
    for question in questions:
        questionId = question[0]
        questionWord = question[1]
        questionMeaning = question[2]
        # 問題文を作成する
        ws[colQuestion + str(curRow)] = questionWord
        # 問題のタイプを指定する
        ws[colQuestionType + str(curRow)] = "Multiple Choice"
        # 選択肢を作成する
        options = handler.execute(
            "SELECT meaning FROM vocabulary WHERE id <> "
            + str(questionId)
            + " ORDER BY RANDOM() LIMIT 5"
        ).fetchall()
        ws[colOption1 + str(curRow)] = options[0][0]
        ws[colOption2 + str(curRow)] = options[1][0]
        ws[colOption3 + str(curRow)] = options[2][0]
        ws[colOption4 + str(curRow)] = options[3][0]
        ws[colOption5 + str(curRow)] = options[4][0]
        # 時間制限を指定する
        ws[colTimeLimit + str(curRow)] = 5
        # 1/2/3/4/5の中で正解を決める
        answerOption = randrange(1, 6)
        ws[colAnswer + str(curRow)] = answerOption
        if answerOption == 1:
            ws[colOption1 + str(curRow)] = questionMeaning
        elif answerOption == 2:
            ws[colOption2 + str(curRow)] = questionMeaning
        elif answerOption == 3:
            ws[colOption3 + str(curRow)] = questionMeaning
        elif answerOption == 4:
            ws[colOption4 + str(curRow)] = questionMeaning
        elif answerOption == 5:
            ws[colOption5 + str(curRow)] = questionMeaning
        else:
            # 何もしない
            pass
        # 画像リンクを空にする
        ws[colImageLink + str(curRow)] = ""
        curRow += 1
    # ファイルを保存する
    wb.save("output/quizizz.xlsx")


def msgbox(message, title="エラー"):
    if platform.system() == "Windows":
        # Windows
        MessageBox = ctypes.windll.user32.MessageBoxW
        MessageBox(None, message, title, 0)
    elif platform.system() == "Darwin":
        # macOS
        os.system(
            "osascript -e 'display alert \"" + title + '" message "' + message + "\"'"
        )
    else:
        # Linux
        # 対応していません
        pass


def open_folder(path):
    if platform.system() == "Windows":
        # Windows
        os.startfile(path)
    elif platform.system() == "Darwin":
        # macOS
        subprocess.Popen(["open", path])
    else:
        # Linux
        subprocess.Popen(["xdg-open", path])


def main(platform):
    # データベースを作成する
    if os.path.exists("sqlite/database.db"):
        os.remove("sqlite/database.db")
    db = sqlite3.connect("sqlite/database.db")
    handler = db.cursor()
    # テーブルを作成する
    handler.execute(
        "CREATE TABLE vocabulary(id INTEGER PRIMARY KEY AUTOINCREMENT, word STRING, meaning STRING)"
    )
    # データをインポートする
    import_data(handler, platform)
    # 変更を反映する
    db.commit()
    # アップロード用ファイルを生成する
    generate_upload_file(handler, platform)
    # データベースの接続を解除する
    db.close()
    # データベースを削除する
    os.remove("sqlite/database.db")


if __name__ == "__main__":

    if os.path.exists("input/_kahoot.xlsx"):
        # Kahoot!
        if os.path.exists("input/kahoot.xlsx"):
            os.remove("input/_kahoot.xlsx")
        else:
            os.rename("input/_kahoot.xlsx", "input/kahoot.xlsx")

    if os.path.exists("input/_quizizz.xlsx"):
        # Quizizz
        if os.path.exists("input/quizizz.xlsx"):
            os.remove("input/_quizizz.xlsx")
        else:
            os.rename("input/_quizizz.xlsx", "input/quizizz.xlsx")

    emptyInput = True
    if os.path.exists("input/kahoot.xlsx"):
        # Kahoot!
        emptyInput = False
        main("Kahoot!")
    if os.path.exists("input/quizizz.xlsx"):
        # Quizizz
        emptyInput = False
        main("Quizizz")
    if emptyInput is True:
        # ファイルが見つかりません
        msgbox(
            "入力フォルダの input/ に kahoot.xlsx または quizizz.xlsx を入れてください。", "ファイルが見つかりません"
        )
        exit(1)
    # 出力フォルダを開く
    open_folder("output")
